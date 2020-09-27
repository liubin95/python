# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from scrapy.utils.project import get_project_settings


class AliyunPipeline(object):
    def open_spider(self, spider):
        """
        开始时调用
        :param spider:
        """
        if os.path.exists('./aliyun.xlsx'):
            os.remove('./aliyun.xlsx')
        spider.logger.info("开始输出xls文件")
        self.wb = Workbook()
        self.page = 0
        self.wb.alignments = Alignment(horizontal='center', vertical='center')
        self.ws = self.wb.create_sheet(title="host", index=0)
        headers = get_project_settings().get('FIELDS_TO_EXPORT')
        self.ws.append(headers)

    def process_item(self, item, spider):
        """
        每一次 返回item时调用
        :param item:
        :param spider:
        :return:
        """
        self.page += 1
        setting = get_project_settings()

        red_fill = PatternFill("solid", fgColor="FF0000")
        write_arr = [item['instance_name'],
                     item['host_name'],
                     item['os_name'],
                     item['instance_status'],
                     item['instance_address'],
                     item['instance_ip'],
                     item['cpu_used'],
                     item['memory_used'],
                     ','.join(item['disk_list'])]
        self.ws.append(write_arr)
        row_num = self.ws.max_row
        # cpu数据
        if (item['cpu_used'] != '无数据') and float(item['cpu_used'].replace('%', '')) >= setting.get('CPU_CORDON'):
            self.ws.cell(row_num, 7).fill = red_fill
        #     内存数据
        if (item['memory_used'] != '无数据') and float(item['memory_used'].replace('%', '')) >= setting.get(
                'MEMORY_CORDON'):
            self.ws.cell(row_num, 8).fill = red_fill
        #     磁盘数据
        for each in item['disk_list']:
            if each.find('(') != -1:
                begin = each.find('(')
                end = each.rfind(')')
                disk_used = each[begin + 1:end]
            else:
                disk_used = each
            if float(disk_used.replace('%', '')) >= setting.get('DISK_CORDON'):
                self.ws.cell(row_num, 9).fill = red_fill

        return item

    def close_spider(self, spider):
        """
        结束时调用
        :param spider:
        """
        spider.logger.info("结束输出xls文件，共%d页".format(self.page))
        self.wb.save(filename='./aliyun.xlsx')
